"""KONSTRYX build tracker — status, decisions, open questions, sequence, issues."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\Ziya\Documents\Claude\Projects\KONSTRYX DEV\KONSTRYX_Status_Tracker.xlsx"
FONT = "Arial"

NAVY = "1F3864"
BLUE = "2E5C8A"
GREY = "F2F2F2"
GREEN = "C6EFCE"
AMBER = "FFEB9C"
RED = "FFC7CE"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def header(ws, row, cols, fill=NAVY):
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 28


def title(ws, text, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name=FONT, bold=True, size=14, color=NAVY)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(row=2, column=2 - 1, value=sub)
    s.font = Font(name=FONT, size=9, italic=True, color="595959")
    ws.row_dimensions[1].height = 20


def body(ws, start_row, rows, widths, status_col=None):
    for r, data in enumerate(rows, start=start_row):
        for i, v in enumerate(data, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.font = Font(name=FONT, size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
        if status_col:
            s = ws.cell(row=r, column=status_col).value
            fill = {"Done": GREEN, "In progress": AMBER, "Not started": None,
                    "Blocked": RED, "At risk": RED, "Open": AMBER,
                    "Decided": GREEN, "Broken": RED, "Closed": GREY,
                    "YES": GREEN, "OPEN": AMBER, "Adopted": GREEN}.get(s)
            if fill:
                ws.cell(row=r, column=status_col).fill = PatternFill("solid", fgColor=fill)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ------------------------------------------------------------------ 1. Status
ws = wb.active
ws.title = "Status Report"
title(ws, "KONSTRYX — Build Status Report",
      "Product: SAP BTP extension for EC&O · CAP Java + SAPUI5 · dedicated deployment per client", 6)

ws["A4"] = "Snapshot"
ws["A4"].font = Font(name=FONT, bold=True, size=11, color=NAVY)

header(ws, 5, ["Metric", "Count", "", "", "", ""])
metrics = [
    ("Work items total", '=COUNTA(\'Work Items\'!B6:B100)'),
    ("Done", '=COUNTIF(\'Work Items\'!E6:E100,"Done")'),
    ("In progress", '=COUNTIF(\'Work Items\'!E6:E100,"In progress")'),
    ("Not started", '=COUNTIF(\'Work Items\'!E6:E100,"Not started")'),
    ("Blocked / at risk", '=COUNTIF(\'Work Items\'!E6:E100,"Blocked")+COUNTIF(\'Work Items\'!E6:E100,"At risk")'),
    ("Decisions taken", '=COUNTA(Decisions!B6:B100)'),
    ("Decisions open (need you)", '=COUNTIF(\'Open Decisions\'!F6:F100,"Open")'),
    ("Known issues open", '=COUNTIF(Issues!F6:F100,"Open")+COUNTIF(Issues!F6:F100,"Broken")'),
    ("Suggestions open", '=COUNTIF(Suggestions!F6:F100,"Open")'),
]
for r, (label, formula) in enumerate(metrics, start=6):
    ws.cell(row=r, column=1, value=label).font = Font(name=FONT, size=10)
    c = ws.cell(row=r, column=2, value=formula)
    c.font = Font(name=FONT, size=10, bold=True)
    c.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=1).border = BORDER
    c.border = BORDER

ws["A16"] = "What is running right now"
ws["A16"].font = Font(name=FONT, bold=True, size=11, color=NAVY)
running = [
    ["CAP Java service", "http://localhost:8090", "8 OData V4 services, H2, seeded, auth enforced"],
    ["UI5 app", "http://localhost:8081", "serve.py proxies /odata to 8090 (same origin)"],
    ["Repository", r"C:\Users\Ziya\Documents\Claude\Projects\KONSTRYX DEV", "git, 48 commits"],
    ["Reference (read-only)", "OneDrive Products/konstrux/Konstrucx", "FTS, wireframe v12, requirements, backlog"],
]
header(ws, 17, ["Component", "Location", "Notes", "", "", ""])
body(ws, 18, running, [34, 46, 62, 20, 20, 20])

# -------------------------------------------------------------- 2. Work Items
ws = wb.create_sheet("Work Items")
title(ws, "Work Items — planned and completed",
      "Status: Done · In progress · Not started · Blocked · At risk", 7)
header(ws, 5, ["ID", "Workstream", "Item", "Detail", "Status", "Evidence / verification", "Next action"])

items = [
    ("P-01", "Platform", "JDK 17 + Maven toolchain", "SapMachine 17.0.20, Maven 3.9.16, JAVA_HOME set at user scope", "Done", "java -version / mvn -v both report expected versions", "—"),
    ("P-02", "Platform", "Build workspace + version control", "Repo outside OneDrive at Documents\\Claude\\Projects\\KONSTRYX DEV; git init; .gitignore", "Done", "9 commits on main", "—"),
    ("P-03", "Platform", "CAP Java service builds and runs", "Fixed 5 blocking defects in the Sprint 0 skeleton", "Done", "All 8 OData V4 services return metadata and data", "—"),
    ("P-04", "Platform", "Local DB + mock personas", "H2 embedded, 5 mock users mirroring xs-security role templates", "Done", "Persona-based 403s observed as designed", "—"),
    ("P-05", "Platform", "Test data separated from product data", "13 fixtures moved db/data -> test/data; both Node and Java config layers set", "Done", "cds build --production emits only currency code list; no PRJ-001/Emaar in output", "—"),
    ("P-07", "Platform", "Delivered content pack mechanism", "Versioned packs applied insert-if-missing, in version order; ships in the jar and readable from ./content; applyContentPacks admin action for upgrades without restart", "Done", "Client edit survived an upgrade pack that contained the same row: 1 inserted, 1 left untouched, client scope preserved", "Reuse for the EC&O starter pack (Q-02)"),
    ("P-06", "Platform", "Document number ranges", "Configurable per object: scope GLOBAL or COMPANY, and pattern, set independently", "Done", "GLOBAL issues RR-2026-0001/RES-2026-0001; COMPANY issues BUD-INFC-2026-0001 and BUD-PMI-2026-0001 advancing independently; numbers issued on draft activation, not draft creation", "—"),
    ("A-01", "Authorization", "Authorization model (S/4 style)", "AuthObject + Activity + Persona + PersonaPermission + UserAssignment scoped by company/project", "Done", "Catalogue: 7 modules, 21 objects, 6 activities", "—"),
    ("A-02", "Authorization", "EffectivePermission view", "Flattens assignments x grants to answer 'what can this person do'", "Done", "23 rows resolved for the 4 seeded personas", "—"),
    ("A-03", "Authorization", "Runtime enforcement handler", "Activity check + company/project instance filtering on every CRUD event", "Done", "Scoped user sees 4 of 5 requests; unrestricted sees 5; own $filter still ANDs; no grant = 403", "—"),
    ("A-04", "Authorization", "Write-path input-data check", "Reject CREATE/UPDATE carrying a project outside the user's scope", "Not started", "—", "Validate after Project Setup exists"),
    ("A-05", "Authorization", "Administration UI", "S/4-like screens to maintain personas, grants, assignments", "Not started", "—", "Part of Masters/Admin UI block"),
    ("F-01", "Frameworks", "Approval framework model", "Schemes per object type, ordered steps, approver persona, value bands, runtime instances", "Done", "Compiles; exposed on /collaboration", "—"),
    ("F-02", "Frameworks", "Approval engine", "Instantiate on submit, match value bands, advance steps, enforce approver persona and separation of duties, delegate, withdraw, audit", "Done", "28 checks against the running service: bands select 1/2/3 steps at 50k/300k/2m; out-of-order, double, reasonless and same-person decisions all refused; persona configured through the admin API then enforced", "Inbox screen (F-04) and per-object wiring (F-05)"),
    ("F-04", "Frameworks", "Approval inbox screen", "The steps awaiting the signed-in user, with approve, reject and delegate", "Not started", "—", "With the UI rebuild"),
    ("F-05", "Frameworks", "Wire submit into each object's lifecycle", "Today the caller chooses when to submit. Each document type should move to 'In Approval' on submit and act on the outcome", "Not started", "—", "Per module as each is built"),
    ("F-06", "Frameworks", "Content pack references", "Packs can name another row instead of its UUID, and match on composite natural keys; each pack applies atomically", "Done", "APPROVAL_SCHEMES resolves auth objects and its own schemes at deploy time: 6 rows inserted", "—"),
    ("F-07", "Frameworks", "Attachments on every object", "One polymorphic attachment table for all modules; automatic versioning with a supersedes chain; client-configurable categories; a mandatory category blocks approval submission", "Done", "14 checks: v1 then v2 superseding it, PDF streamed in and read back byte-identical, 404 for a missing target, and submission refused until the mandatory drawing was attached", "fileSize (F-08) and the UI upload control"),
    ("F-08", "Frameworks", "Attachment file size", "fileSize is null — measuring the media stream consumes it, so it needs a counting wrapper on the upload path", "Not started", "—", "Low priority; display only"),
    ("F-10", "Frameworks", "Per-user table personalization", "Saved column sets, filters, sorts and groupings per user per table; one default each; administrator-published variants readable by all", "Done", "18 checks: two users hold the same variant name on the same table and see only their own; a new default clears the previous one and leaves other users alone; writing to someone else's answers 404, not 403, so the key is not confirmed", "UI wiring comes with the screen rebuild"),
    ("F-09", "Frameworks", "Attachment storage", "Content sits in the database as LargeBinary. Fine for drawings and permits; a project's photo library is a different question", "Not started", "—", "Decide before go-live — see S-24"),
    ("F-03", "Frameworks", "Attachments model", "Polymorphic on entityName+objectID, media-type content, categories", "Done", "Compiles; exposed on /collaboration", "Upload handler + object store not built"),
    ("F-04", "Frameworks", "Attachment upload + storage", "Object store binding for Cloud Foundry vs HANA LOB", "Not started", "—", "Decision needed (see Open Decisions)"),
    ("F-05", "Frameworks", "Table personalization store", "UserVariant entity holding UI5 p13n state per user", "Done", "Compiles; exposed on /collaboration", "Not wired to any table yet"),
    ("F-06", "Frameworks", "Table personalization UI wiring", "p13n + VariantManagement on every list screen", "Not started", "—", "Apply as each screen is built"),
    ("M-01", "Data model", "Workflow spine RR->ADV->AVC->RES", "Vertical-agnostic, multi-line, seeded with canonical EQR thread", "Done", "Queryable over OData; 5 lines, AED 685,080", "—"),
    ("M-02", "Data model", "WBS on request line", "Spec section 6 requires one WBS per line; was missing", "Done", "3 distinct WBS across the 5 EQR lines", "—"),
    ("M-03", "Data model", "EQR vertical extension", "Instances, mob/demob window, own-vs-rental, vendor, operators", "Done", "$expand=equipment,wbs returns all 5 lines resolved", "—"),
    ("M-04", "Data model", "Chain steps CMT/MOB/OPL/VAR/DMB/CLS", "Steps 5-10 have no CDS entities at all", "Not started", "—", "Part of Project Execution block"),
    ("U-01", "UI", "RR worklist on live OData", "RequestOverview projection, server-side filters", "Done", "Verified in browser: 4 requests, EQR shows 5 lines / 685,080; EQR filter issues new $batch", "—"),
    ("U-02", "UI", "Request detail page on OData", "Still reads webapp/model/data.json", "Not started", "—", "Now unblocked by M-03"),
    ("U-03", "UI", "Chain step pages on OData", "Still read data.json", "Blocked", "—", "Blocked by M-04"),
    ("U-04", "UI", "Launchpad intents declared", "crossNavigation inbounds for KonstryxResourceRequest and KonstryxReservation", "Done", "manifest parses; two inbounds registered", "—"),
    ("U-05", "UI", "Launchpad-hosted shell (blend with S/4)", "App shell removed: root view is now the NavContainer alone. Launchpad supplies header, search, user menu, theme", "Done", "Verified in browser as daud: worklist renders with no app chrome, 4 of 5 requests correctly scoped, EQR at AED 685,080", "—"),
    ("U-06", "UI", "Spaces and Pages navigation", "flp.html + flpSite.json written: CDM 3.1 site, one space, one page, two intent tiles", "At risk", "Site JSON authored against the runtime's own sandboxSite.json schema", "Sandbox bootstrap fails reading a null script element; finish or verify on real BTP instead"),
    ("U-08", "UI", "Split into task-focused apps per document type", "S/4 pattern is one app per task, reached from a tile. Currently one app with in-app routing across ten document types", "Not started", "—", "Register more intents as screens are built"),
    ("U-07", "UI", "Morning Horizon theme", "sap_horizon is Morning Horizon; already the bootstrap theme", "Done", "index.html bootstraps data-sap-ui-theme=sap_horizon", "Must be inherited from the shell once launchpad-hosted, never hard-coded"),
    ("B-01a", "Business", "Masters — hybrid scope enforcement", "Reads on any entity carrying the scoped aspect narrow to GROUP or the user's own companies; aspect duck-typed from the model, not a name list", "Done", "Two stewards in different legal entities each see 5 group masters + their own local one, and neither sees the other's", "—"),
    ("B-01b", "Business", "Masters — promotion queue (mdg)", "requestPromotion on the master, approve/reject on the queue; scope is never edited directly", "Done", "EQ-LOC-INFC invisible to PMI, requested, approved, then visible — reason, decider and outcome recorded", "—"),
    ("B-01c", "Business", "Masters — validation rules", "Level/parent agreement, code uniqueness per scope, rate effective-date clashes, promotion collision check", "Done", "Valid L1 accepted; L5 with no parent, L1 given a parent, L3 under an L5, duplicate code and group-vs-local clash all rejected with 409 and a specific message; promotion blocked while INFC and PMI both hold EQ-DUP, naming both", "—"),
    ("B-01d", "Business", "Masters — remaining entities catalogued", "CBS library, productivity and consumption norms, templates, material mirror added to the authorization catalogue", "Done", "CBS library picked up isolation, promotion and L1-L3 depth checks with no CBS-specific code — both handlers duck-type the scoped aspect", "—"),
    ("B-01e", "Business", "Masters — list screens", "Resource hierarchy, CBS library, rate master and promotion queue, all on live OData with server-side filters", "Done", "Verified by screenshot: 7 resources (PMI's local one correctly absent from an INFC session), 7 CBS nodes, 8 rates, 1 pending promotion", "—"),
    ("B-01f", "Business", "Masters — object page", "Header facts, attributes, rate history and hierarchy children; route carries the code, not the UUID", "Done", "EQ-TWC-12T shows all three of its rates together; EQ-TOWER shows parent EQ-CRANE and child EQ-TWC-12T", "—"),
    ("B-01h", "Business", "Masters — draft editing", "CAP draft: private copy on edit, stored record untouched until activation, resumes an abandoned draft", "Done", "Edit that was previously rejected now activates and persists; validation still fires on activation", "Edit is on the resource page only; the other masters reuse the same pattern"),
    ("B-01g", "Business", "Masters — productivity and consumption norms", "One screen, two tabs; plus a material branch in the resource tree so consumption norms attach to materials rather than equipment", "Done", "4 productivity norms with crew composition, 3 consumption norms with wastage; group 105 kg/m3 rebar against INFC's own 112 kg for coastal detailing", "—"),
    ("B-02a", "Business", "Templates — model and instantiation", "Template carries construction type, CBS structure and default resources; instantiate copies them into a project, two-pass so children parent to their new instances", "Done", "TPL-HIGHRISE into PRJ-002: 7 CBS nodes (3 roots, 4 correctly parented, all traced to library) and 6 planned resources; second run refused", "—"),
    ("B-02b", "Business", "Templates — screen", "List plus an instantiate dialog offering the projects the user may see; refusal surfaces the service's own message", "Done", "Driven through the UI: TPL-HIGHRISE into PRJ-002 created 7 CBS nodes and 6 planned resources; a second run raised the refusal", "Object page for a template still to do"),
    ("B-03a", "Business", "Project Setup", "Project master, WBS, CBS instance; S/4 Enterprise Project mirror", "Not started", "—", "Next block; S/4 tenant (Q-09) becomes the blocker for the mirror"),
    ("B-02", "Business", "Templates", "Project templates: CBS tree + default resources", "Not started", "—", "After Masters"),
    ("B-03", "Business", "Project Setup", "Project, WBS, CBS instance; S/4 Enterprise Project mirror", "Not started", "—", "After Templates"),
    ("B-04", "Business", "Uploads", "BOQ import, estimate import, versioning with progress carry-forward", "Not started", "—", "Highest-risk item in the sequence"),
    ("B-05", "Business", "Project Planning", "Activity layer, native entity + P6 adapter", "Not started", "—", "After Uploads"),
    ("B-06", "Business", "Budgeting", "Budget from BOQ/CBS, 4-category ledger, approval, baseline, encumbrance", "Not started", "—", "After Planning"),
    ("B-07", "Business", "Project Execution", "RR->CLS chain end to end, material + manpower", "Not started", "—", "After Budgeting"),
    ("B-08", "Business", "Project Commercial", "Client billing, variations, payment certificates", "Not started", "—", "After Execution"),
    ("B-09", "Business", "Plant Department", "Equipment, fleet, scaffolding/formwork", "Not started", "—", "Last per your sequence"),
    ("D-01", "Deployment", "MTA builds a deployable archive", "Approuter + srv + db-deployer; XSUAA, HANA, Destination resources", "Done", "konstryx_0.1.0.mtar 68.9 MB; 234 HDI artifacts and .hdiconfig verified inside the db-deployer", "—"),
    ("D-04", "Deployment", "First Cloud Foundry deployment", "Push the archive, bind services, verify the app end to end in the cloud", "Blocked", "—", "Blocked by I-24: needs a KONSTRYX subaccount and cf login"),
    ("D-02", "Deployment", "S/4HANA Public Cloud connector", "Mirror entities exist; no connector, no destination wiring", "Not started", "—", "Needs S/4 dev tenant access"),
    ("D-03", "Deployment", "CI/CD pipeline", "Not started", "Not started", "—", "Decision needed"),
]
body(ws, 6, items, [8, 15, 30, 52, 13, 56, 30], status_col=5)
ws.freeze_panes = "A6"

# --------------------------------------------------------------- 3. Decisions
ws = wb.create_sheet("Decisions")
title(ws, "Decisions taken — for your review",
      "Everything here is currently binding. Tell me if any should be reopened.", 7)
header(ws, 5, ["ID", "Date", "Decision", "Rationale", "Decided by", "Impact if reversed", "Status"])

decisions = [
    ("D-01", "2026-08-15", "Runtime is CAP Java, not Node.js", "Skeleton, pom.xml, mta.yaml and generated EDMX are already Java; phase plan staffs 2 CAP Java devs; workload is integration-heavy and long-lived", "Claude (delegated by you)", "High — rewrites pom, mta, handlers and the staffing plan", "Decided"),
    ("D-02", "2026-08-15", "Dedicated deployment per client, not shared multitenancy", "Your instruction. tenant-mode stays dedicated; no MTX sidecar, no SaaS registry", "You", "High — MTX retrofit, HDI-per-tenant, subscription callbacks", "Decided"),
    ("D-03", "2026-08-15", "Code lists ship; business content is an optional import", "Your instruction. db/data holds only what every client legitimately receives", "You", "Medium — changes onboarding and what db/data contains", "Decided"),
    ("D-04", "2026-08-15", "Vertical-specific extension entity per request line", "Your instruction. Six verticals differ genuinely; one flat line entity would be mostly null", "You", "Medium — reshapes all six verticals", "Decided"),
    ("D-05", "2026-08-15", "Evolve the existing freestyle UI5 app to OData V4", "Preserves the app stakeholders already validated as 'how it will look when developed'", "You", "High — discards the validated app", "Decided"),
    ("D-06", "2026-08-15", "First increment is a vertical slice, not platform-first", "Proves CDS + handlers + auth + UI binding end to end before scaling", "You", "Low — sequencing only", "Decided"),
    ("D-07", "2026-08-15", "Build repo outside OneDrive, git-tracked", "OneDrive sync on node_modules/target/gen causes file locks and corrupted builds", "You", "Low", "Decided"),
    ("D-08", "2026-08-15", "CAP Java 4.9.3 + Spring Boot 3.5.6", "CAP Java 3.5.0 rejects cds-compiler 6 shipped by @sap/cds 9; every query returned 500", "Claude (technical)", "Low — but do not drop to 3.x without pinning cds-dk to 8", "Decided"),
    ("D-09", "2026-08-15", "Authorization is configured at runtime, not compiled into XSUAA scopes", "Your requirement for S/4-style in-app administration by project, company and module", "You", "High — the whole auth layer", "Decided"),
    ("D-10", "2026-08-15", "XSUAA Admin role bypasses the data-driven auth layer", "A fresh deployment has no personas or assignments; without it nobody could sign in to create the first one", "Claude (technical)", "Low — but it is a standing privileged path, worth a security review", "Decided"),
    ("D-11", "2026-08-15", "Seeded EQR line values at AED 685,080, not the 716,044 header", "The line values are self-consistent and reconcile exactly to the per-WBS commitment breakdown; the header does not", "Claude (flagged to you)", "Low technically, but the demo numbers must be settled", "Decided"),
    ("D-12", "2026-08-15", "CAP service listens on 8090 locally", "8080 is taken by the UI5 dev server", "Claude (technical)", "Trivial", "Decided"),
    ("D-13", "2026-08-15", "Number range scope is configurable per object, not a product-wide choice", "Your instruction. Scope (GLOBAL/COMPANY) and pattern are configured independently so a client can print the company code while running one group series, or vice versa", "You", "Low — configuration only", "Decided"),
    ("D-14", "2026-08-15", "Document numbers are issued on draft activation, not draft creation", "Abandoned drafts would otherwise burn numbers and leave gaps in the series", "Claude (technical)", "Low, but gap-free numbering is sometimes an audit requirement — confirm if so", "Decided"),
    ("D-15", "2026-08-15", "UI is launchpad-hosted; Spaces and Pages are launchpad configuration, not application data", "Your requirement to blend with S/4HANA. The shell, theme and navigation come from the launchpad; the app renders content only. Modelling spaces as CDS entities was started and abandoned as wrong for this target", "You", "High — supersedes part of D-05; the app's own ToolPage shell must be removed", "Decided"),
    ("D-16", "2026-08-15", "Option A — S/4HANA Public Cloud's own launchpad is the shell, not SAP Build Work Zone", "Users live in S/4 all day, so one shell is the strongest blend. Avoids a Work Zone dependency and licence per client. App registered as IAM External App via LADI -> Business Catalog -> Business Role", "You", "Medium — Work Zone remains available later; app-side work (intents, no own shell) is identical either way, so switching costs launchpad configuration only", "Decided"),
    ("D-17", "2026-08-15", "The project is mastered in KONSTRYX, not in S/4", "Your instruction. A project is created in KONSTRYX and synchronised outward to S/4, or brought in from Primavera P6. This reverses the KONSTRYX-as-reader assumption for the project object specifically; procurement and finance stay S/4-owned", "You", "High — prj.Project is currently @readonly as an S/4 mirror and must become writable with outbound sync and an unsynchronised state visible in the UI", "Decided"),
    ("D-18", "2026-08-15", "Every master and transaction gets upload and download", "Your instruction. Excel import/export as a first-class capability rather than per-screen bespoke work, and the route by which projects arrive from P6", "You", "High — a shared import/export framework touching every module; needs template definition, validation, error reporting and a staging area", "Decided"),
    ("D-19", "2026-08-15", "Flexible approval workflow on every object", "Your instruction. Extends the approval framework already modelled from budgets to all masters and transactions, with the workflow configurable rather than coded", "You", "Engine built and verified. Value bands, approver personas, separation of duties, delegation and withdrawal are all configuration. Remaining: the inbox screen and wiring submit into each object's lifecycle", "Decided — engine done"),
]
body(ws, 6, decisions, [8, 12, 44, 62, 22, 42, 12], status_col=7)
ws.freeze_panes = "A6"

# ---------------------------------------------------------- 4. Open Decisions
ws = wb.create_sheet("Open Decisions")
title(ws, "Open decisions — needed from you",
      "These block or shape work already in the sequence. Ordered by when I need the answer.", 7)
header(ws, 5, ["ID", "Question", "Why it matters", "Options", "Blocks", "Status", "Needed by"])

opens = [
    ("Q-01", "Document number ranges: per company or group-global?", "ANSWERED: configurable per object. Built and verified both scopes", "—", "—", "Closed", "Answered 2026-08-15"),
    ("Q-13", "Where does KONSTRYX surface — Work Zone or the S/4 launchpad?", "ANSWERED: Option A, S/4's own launchpad. See D-16", "—", "—", "Closed", "Answered 2026-08-15"),
    ("Q-14", "Does numbering need to be gap-free for audit?", "Numbers are issued on activation so abandoned drafts do not consume them, but a cancelled document still leaves a gap. Some jurisdictions require unbroken sequences for certain document types", "Gaps acceptable (current) / gap-free required for named document types", "P-06", "Open", "Before Budgeting"),
    ("Q-02", "What is in the EC&O starter content pack?", "You chose 'code lists ship, starter pack as optional import'. I need to know what the pack contains before building the import", "Resource hierarchy depth, CBS library, productivity/consumption norms, trade catalogue", "B-01 Masters, B-02 Templates", "Open", "Before Masters"),
    ("Q-03", "Approval schemes: who approves what, at which value bands?", "The engine is built and two starting schemes ship with it — Budget (100k / 1m) and Resource Request (250k / 1m), both with no approver persona set because personas are yours to name. Those thresholds are a placeholder I invented; they are not your policy", "Confirm or replace the thresholds, and name the persona that approves each step. All of it is configurable in the app — no redeploy", "F-02 is done and running on these defaults", "Open", "Before the first client demo"),
    ("Q-04", "Attachment storage: SAP Object Store or HANA LOB?", "Object Store needs a CF service instance and changes the upload path; HANA LOB is simpler but costlier at volume", "Object Store (recommended for drawings/photos) / HANA LOB", "F-04", "Open", "Before Uploads"),
    ("Q-05", "The canonical demo numbers do not reconcile", "Wireframe header says AED 716,044; the five line values sum to 685,080 and match the per-WBS commitment split exactly. L1 rate 320/day implies 115,200 not 276,480", "Correct the header to 685,080 / correct L1 rate to 768 / supply the intended figures", "Any stakeholder demo", "Open", "Before first demo"),
    ("Q-06", "BOQ import template — confirm the column set", "Named the highest-risk item in your own requirements register. The importer is built to the template, not the reverse", "Provide the actual client BOQ template(s) you must accept", "B-04 Uploads", "Open", "Before Uploads"),
    ("Q-07", "CBS instance versioning: copy-on-create or live reference to the library?", "Open item in Data Model Spec section 10. Affects whether library changes propagate into running projects", "Copy-on-create (proposed) / live reference", "B-01, B-03", "Open", "Before Project Setup"),
    ("Q-08", "Encumbrance currency: company currency or group reporting currency?", "Open item in Data Model Spec section 10", "Company ccy + group conversion in ins (proposed) / group ccy", "B-06 Budgeting", "Open", "Before Budgeting"),
    ("Q-09", "S/4 dev tenant — what is actually needed (merges the old S-15)", "Nothing S/4-facing can be built or tested without it. Four concrete things: (1) the tenant API host, e.g. https://myNNNNNN-api.s4hana.cloud.sap; (2) a communication user with password, or a client certificate; (3) communication arrangements activated - SAP_COM_0308 first, since the project now syncs OUTWARD; (4) a named S/4 administrator who can set those up and re-activate them when they expire", "Provide all four, or confirm the connector is deferred", "Project sync (D-17), all S/4 mirrors", "Open", "Before the S/4 connector - NOT blocking upload/download or approvals"),
    ("Q-10", "Scope re-baseline: personalization + attachments + approvals on every screen", "Wireframe v12 is 41 modules / 434 screens. The 22-week MVP plan with 1 UI5 developer does not carry this", "Re-baseline the plan / reduce MVP screen scope / add UI capacity", "Overall plan credibility", "Open", "Before committing a client date"),
    ("Q-11", "UI5 production delivery: how is the app served in Cloud Foundry?", "mta.yaml has no approuter and no UI module today, and the app bootstraps from a local runtime path", "Approuter + HTML5 repo (standard) / other", "D-01", "Open", "Before first CF deploy"),
    ("Q-12", "CI/CD tooling", "Prerequisite 4 in the phase plan, still undecided", "SAP CI/CD service / GitHub Actions / Azure DevOps", "D-03", "Open", "Before team scales up"),
]
body(ws, 6, opens, [8, 56, 58, 50, 26, 10, 22], status_col=6)
ws.freeze_panes = "A6"

# ----------------------------------------------------------------- 5. Sequence
ws = wb.create_sheet("Planned Sequence")
title(ws, "Planned build sequence",
      "Your order, with two platform items inserted first because every module inherits them", 6)
header(ws, 5, ["#", "Block", "Contents", "Why here", "Depends on", "Status"])

seq = [
    (1, "Authorization enforcement", "Runtime handler applying configured grants + instance filtering", "Every screen inherits it; retrofitting enforcement after screens exist is where products go wrong", "—", "Done"),
    (2, "Document number ranges", "Configurable scope (GLOBAL/COMPANY) and pattern per object; issued on activation", "Every document module needs it; changing the scheme later means renumbering live data", "Q-01 (answered)", "Done"),
    ("2b", "Launchpad-hosted shell", "Strip the app's own ToolPage chrome; FLP sandbox with Spaces and Pages for local dev", "Blend with S/4HANA. Doing this before the screen count grows avoids stripping the shell out of every screen later", "Q-13", "Not started"),
    (3, "Masters", "Resource hierarchy, CBS library, rates, vendors, scoped GROUP/COMPANY + promotion queue", "Your sequence. Everything downstream references master codes", "Q-01, Q-02", "Not started"),
    (4, "Templates", "Project templates: CBS tree + default resources", "Your sequence. Templates are composed of masters", "Masters", "Not started"),
    (5, "Project Setup", "Project, WBS, CBS instance, S/4 Enterprise Project mirror", "Your sequence. Instantiates a template", "Templates, Q-07, Q-09", "Not started"),
    (6, "Uploads", "BOQ import, estimate import, versioning with progress carry-forward", "Your sequence. Needs a project to import into. Highest-risk item in the plan", "Project Setup, Q-04, Q-06", "Not started"),
    (7, "Project Planning", "Activity layer, native entity + P6 adapter", "Your sequence. Activities hang off WBS", "Project Setup", "Not started"),
    (8, "Budgeting", "Budget from BOQ/CBS, 4-category ledger, approval, baseline, encumbrance", "Your sequence. Needs BOQ and CBS to exist", "Uploads, Q-03, Q-08", "Not started"),
    (9, "Project Execution", "RR->CLS chain end to end, material + manpower, chain steps 5-10", "Your sequence. Consumes budget; encumbrance must exist first", "Budgeting", "Not started"),
    (10, "Project Commercial", "Client billing, variations, payment certificates", "Your sequence. Bills against executed work", "Execution", "Not started"),
    (11, "Plant Department", "Equipment, fleet, scaffolding/formwork", "Your sequence — explicitly after the core", "Execution", "Not started"),
]
body(ws, 6, seq, [5, 30, 62, 66, 26, 13], status_col=6)
ws.freeze_panes = "A6"

# ------------------------------------------------------ 5b. Suggestions & flags
ws = wb.create_sheet("Suggestions")
title(ws, "Suggestions and flags raised during the build",
      "Everything recommended or warned about, kept in one place for review. Nothing here is decided.", 7)
header(ws, 5, ["ID", "Area", "Suggestion or flag", "Why it matters", "My recommendation", "Status", "Raised"])

suggestions = [
    ("S-01", "Plan", "Re-baseline the MVP estimate", "Personalization, attachments and approvals apply to every object. Wireframe v12 is 41 modules / 434 screens; the 22-week plan with one UI5 developer does not carry that layer", "Re-baseline before any client date is committed, or cut MVP screen scope", "Open", "2026-08-15"),
    ("S-02", "Architecture", "Revisit tenancy once client count grows", "Dedicated deployment means N upgrades for N clients. The model is tenancy-neutral so the switch is cheap now and costly per live client later", "Revisit at 5-10 clients; decide before the first go-live, not after", "Open", "2026-08-15"),
    ("S-03", "UI", "Split into task-focused apps per document type", "S/4 surfaces one app per task, reached from a tile. KONSTRYX is currently one app with in-app routing across ten document types", "Split as screens are built; register an intent per app", "Open", "2026-08-15"),
    ("S-04", "Deployment", "Move the UI to the HTML5 application repository", "The approuter currently routes the UI through the Java service, which is fine for one app but wrong for scaling and caching", "Move when the app splits into several apps", "Open", "2026-08-15"),
    ("S-05", "Authorization", "Validate declared auth paths at startup", "A path that does not resolve on a projection fails per request with a 500 instead of at boot. That is exactly how the RequestOverview outage happened", "Add a startup check over the catalogue against the model", "Open", "2026-08-15"),
    ("S-06", "Local dev", "H2 is in-memory - nothing survives a restart", "Promotion requests, drafts and anything created during a demo are lost when the service restarts. Seed fixtures reload; transactional work does not", "Make H2 file-backed, or point local dev at HANA Cloud, before any live demo", "Open", "2026-08-15"),
    ("S-07", "Build", "Run npm install after every mbt build", "The MTA build prunes devDependencies at the root, removing @sap/cds-dk, and the next Maven build fails with 'cds' is not recognized", "Add it to the build script or CI pipeline", "Open", "2026-08-15"),
    ("S-08", "Quality", "Verify UI work visually, not through the DOM", "DOM text, network traces and per-element geometry all reported success while the screen was blank for an hour. Only a screenshot and an ancestor-chain walk found it", "Screenshot every UI change; measure ancestors, not the element", "Adopted", "2026-08-15"),
    ("S-09", "Demo data", "Settle the canonical figures", "The wireframe header says AED 716,044 while its own line values sum to 685,080 and reconcile to the per-WBS commitment split", "Correct the header, or supply the intended numbers", "Open", "2026-08-15"),
    ("S-10", "Compliance", "Confirm whether numbering must be gap-free", "Numbers are issued on activation so abandoned drafts do not consume them, but a cancelled document still leaves a gap", "Confirm per document type; some jurisdictions require unbroken sequences", "Open", "2026-08-15"),
    ("S-11", "Masters", "Promotion collision currently blocks", "When two companies hold the same code locally, promotion is refused and names both. Merge or force-rename are the alternatives", "Keep blocking unless the business prefers a merge; it is a one-method change", "Open", "2026-08-15"),
    ("S-12", "Security", "Review role collections before go-live", "The XSUAA Admin role bypasses the data-driven authorization layer as a bootstrap path, and service-level @requires still gates entry independently of the persona model", "Security review of both layers before the first client", "Open", "2026-08-15"),
    ("S-13", "Masters", "Reuse the content pack mechanism for the starter pack", "Versioned, insert-if-missing, never overwrites client edits - already built and proven for number ranges", "Ship the EC&O starter pack the same way", "Open", "2026-08-15"),
    ("S-14", "Masters", "Master editing needs draft handling", "A master screen that cannot be maintained is half a screen, and the draft pattern is inherited by every later module", "Build it before Templates so the pattern is settled once", "In progress", "2026-08-15"),
    ("S-15", "Integration", "Merged into Q-09", "Was a duplicate of the S/4 tenant question; the concrete list of what is needed now lives on Q-09", "See Q-09", "Merged", "2026-08-15"),
    ("S-16", "Architecture", "The reader principle now has an exception", "D-17 makes the project KONSTRYX-mastered while procurement and finance stay S/4-owned. The principle is no longer 'S/4 owns transactions'; it is object by object", "Restate the rule in the product documentation so the exception is deliberate rather than remembered", "Open", "2026-08-15"),
    ("S-17", "Integration", "An unsynchronised project is dangerous", "A project created in KONSTRYX that failed to reach S/4 will still accept requests and budgets that can never post", "Show the unsynchronised state prominently in the UI and block budget release until the S/4 project exists", "Open", "2026-08-15"),
    ("S-18", "Architecture", "Build upload/download once, not per screen", "D-18 applies to every master and transaction. Bespoke import per screen is how 40 modules end up with 40 different error behaviours", "One framework: template definition, staging, validation against the same service rules, and an error report the user can correct and re-upload", "Done", "2026-08-15"),
    ("S-19", "Approvals", "The delivered value bands are mine, not yours", "The engine ships with Budget at 100k/1m and Resource Request at 250k/1m so it works out of the box. Nobody at Inflexion or a client has agreed those figures, and a demo will show them as if they were policy", "Replace them before any client sees the product. See Q-03", "Open", "2026-08-15"),
    ("S-20", "Approvals", "Separation of duties is on by default", "Someone who cleared step 1 cannot clear step 2 of the same document. In a small contractor one director genuinely is both signatures, so allowChaining exists per step to permit it", "Confirm the default is right for EC&O clients; it is a per-step switch either way", "Open", "2026-08-15"),
    ("S-21", "Approvals", "Nothing yet withdraws an approval when the document changes", "A resource request approved at AED 2m stays approved if someone later edits it down to 200k, or up. The approval records the amount it was judged on but nothing re-checks it", "Decide per object type whether an edit invalidates an in-flight or completed approval; wire it with F-05", "Open", "2026-08-15"),
    ("S-22", "Approvals", "Editing a scheme mid-flight", "Instances freeze the step number and name at submission, so an in-flight approval survives a scheme edit. It still points at the step definition for the approver persona, and CAP draft activation can renumber those rows", "Confirm the intended behaviour: should a scheme edit affect approvals already running? Today it partly does", "Open", "2026-08-15"),
    ("S-24", "Attachments", "Attachment content is stored in the database", "LargeBinary in HANA is right for the documents that carry legal weight — drawings, permits, signed variations — because they are backed up and restored with the data they belong to. It is the wrong home for thousands of site photographs", "Decide whether photo-heavy objects go to an object store instead. It is a per-category switch if decided before volume builds up, and a migration afterwards", "Open", "2026-08-15"),
    ("S-25", "Attachments", "Nothing scans uploads or limits their size", "Any authenticated user can upload any file of any size to any object. On a client tenant that is both a malware route and a storage risk", "Add a size cap and virus scanning before the first client upload. SAP BTP has a Malware Scanning service; the size cap is configuration", "Open", "2026-08-15"),
    ("S-26", "Attachments", "Deleting an attachment breaks the version chain", "supersedes points at the previous version. Deleting a middle version leaves the chain dangling, and deleting the document someone approved against destroys the evidence", "Block deletion of a superseded version, or of anything attached to a closed approval; mark as obsolete instead", "Open", "2026-08-15"),
    ("S-23", "Frameworks", "Content packs now carry references", "Delivered content could previously only hold flat rows, which is why the approval schemes could not ship as content. Packs now resolve a row by natural key at deploy time and match on composite keys", "Use the same mechanism for the EC&O starter pack (S-13) and for delivered personas", "Open", "2026-08-15"),
]
body(ws, 6, suggestions, [8, 16, 44, 62, 52, 12, 12], status_col=6)
ws.freeze_panes = "A6"

# --------------------------------------------------------- 6. Options register
ws = wb.create_sheet("Options Considered")
title(ws, "Options register — what was on the table, and what switching would cost",
      "Kept so any decision can be reopened without re-deriving the alternatives. Chosen option marked.", 7)
header(ws, 5, ["Decision", "Question", "Option", "Chosen", "Trade-off", "Cost to switch to this later", "Reversibility"])

options = [
    ("D-01", "Backend runtime", "CAP Java", "YES", "Matches skeleton, staffing plan and integration-heavy workload; slower inner loop", "—", "Hard — rewrites pom, mta, handlers"),
    ("D-01", "", "CAP Node.js", "", "Fastest iteration, lower memory; contradicts every signed plan document", "Rewrite all service handlers and deployment descriptors", "Hard"),
    ("D-02", "Tenancy", "Dedicated — one deployment per client", "YES", "Simplest to build and certify; upgrade effort scales linearly with clients", "—", "Reversible now, costly per live client later"),
    ("D-02", "", "Shared multitenancy", "", "One upgrade serves all tenants; needs MTX sidecar, SaaS registry, subscription callbacks", "A few days now with zero clients. Per live client later: data copy into a tenant container, freeze window, rollback plan", "Model is tenancy-neutral, so no data reshaping"),
    ("D-03", "Delivered content", "Code lists ship; business content is an optional import", "YES", "Clean tenant databases, content versioned independently", "—", "Easy"),
    ("D-03", "", "Ship a starter EC&O pack in the database", "", "Faster onboarding; becomes product IP you must version and upgrade across tenants", "Move rows from the import into db/data", "Easy"),
    ("D-03", "", "Code lists only, nothing else ever", "", "Purest no-assumptions reading; slowest client onboarding", "Drop the import mechanism", "Easy"),
    ("D-04", "Vertical modelling", "Extension entity per vertical", "YES", "Fits six genuinely different verticals; more entities, polymorphic queries", "—", "Medium"),
    ("D-04", "", "Denormalise onto the request line", "", "Cheaper queries; routing decision then exists in two places and can drift", "Flatten extension entities into the spine", "Medium"),
    ("D-04", "", "Derive routing from AdvisoryDecision", "", "No duplication; cannot express re-sourcing by variation without reopening ADV", "Drop extension entities, add fields to ADV", "Medium"),
    ("D-16", "Launchpad / shell", "Option A — S/4's own Fiori launchpad", "YES", "Strongest blend for users who live in S/4; no Work Zone licence or dependency per client", "—", "Easy — app-side work is identical either way"),
    ("D-16", "", "Option B — SAP Build Work Zone", "", "One entry point across S/4 + SuccessFactors + Ariba + extensions; entitled at 1 user per FUE but another moving part per client", "Launchpad configuration only; no application change", "Easy"),
    ("D-16", "", "Standalone app with its own shell", "", "What exists today. Does not blend — duplicates launchpad chrome", "Keep the ToolPage shell", "Easy"),
    ("D-13", "Number ranges", "Configurable scope and pattern per object", "YES", "Serves both client styles from one build", "—", "Easy"),
    ("D-13", "", "Fixed per company, or fixed global", "", "Simpler; forces one convention on every client", "Remove the scope setting", "Easy"),
    ("D-05", "UI approach", "Evolve the existing freestyle UI5 app", "YES", "Keeps the app stakeholders validated; manual binding work per screen", "—", "Medium"),
    ("D-05", "", "Fiori Elements throughout", "", "Fastest across 434 screens; chain strip, wizards and mass-entry grids do not fit the floorplans", "Regenerate screens from annotations", "Medium"),
    ("D-05", "", "Hybrid — FE for masters, freestyle for chain", "", "Best effort/fidelity ratio at scale; two patterns to maintain", "Adopt incrementally, per module", "Easy"),
    ("Q-04", "Attachment storage", "SAP Object Store", "OPEN", "Right for drawings and photos at volume; needs a CF service instance", "—", "Medium once files exist"),
    ("Q-04", "", "HANA LOB", "OPEN", "Simpler, no extra service; costlier at volume and bloats the container", "—", "Medium once files exist"),
]
body(ws, 6, options, [10, 22, 44, 9, 60, 56, 34], status_col=4)
ws.freeze_panes = "A6"

# ------------------------------------------------------------------ 7. Issues
ws = wb.create_sheet("Issues")
title(ws, "Issues, risks and things that broke",
      "Defects found and fixed are listed too, so the history is visible", 7)
header(ws, 5, ["ID", "Area", "What happened", "Impact", "Resolution", "Status", "Raised"])

issues = [
    ("I-01", "CAP build", "BudgetServiceHandler imported com.sap.cds.services.cds.CdsService, which does not exist", "Service would not compile", "Replaced with EventContext", "Closed", "2026-08-15"),
    ("I-02", "CAP build", "pom used cds-starter-spring-boot, which carries no protocol adapter", "All 6 services registered but exposed ZERO HTTP endpoints, silently", "Switched to cds-starter-spring-boot-odata", "Closed", "2026-08-15"),
    ("I-03", "CAP build", ".cdsrc.json set build.target=gen (Node layout)", "Compiled model never reached srv resources; runtime started with an empty catalogue", "Removed; Java build now writes in place", "Closed", "2026-08-15"),
    ("I-04", "CAP build", "CAP Java 3.5.0 rejects cds-compiler 6 from @sap/cds 9", "Every query returned HTTP 500", "Upgraded to cds-services 4.9.3 / Spring Boot 3.5.6", "Closed", "2026-08-15"),
    ("I-05", "Deployment", "mta.yaml declared build-result target/*-exec.jar but build produced konstryx-srv.jar", "A Cloud Foundry deploy would have failed", "Added the exec classifier", "Closed", "2026-08-15"),
    ("I-06", "Deployment", "mta.yaml db-deployer pointed at gen/db, which the Java build does not produce", "HDI deploy would have failed", "Corrected path to db/", "Closed", "2026-08-15"),
    ("I-07", "Data safety", "Demo fixtures were in db/data, which CAP deploys to every environment", "PRJ-001 Marina Heights would have landed in client databases", "Moved 13 fixtures to test/data; both config layers set; cloud profile pins initialization-mode never", "Closed", "2026-08-15"),
    ("I-08", "UI", "Proxy dropped the OData-MaxVersion header", "CAP answered OData-Version 4.01; UI5 V4 model rejected the batch, table silently empty", "Proxy now forwards headers wholesale rather than by allowlist", "Closed", "2026-08-15"),
    ("I-09", "Authorization", "Handler matched the service projection name against the persistence entity name (plural vs singular)", "Handler silently protected NOTHING. Early 403s came from pre-existing service-level @requires, not from the handler", "Resolves the projection to its source entity via CDS model reflection", "Closed", "2026-08-15"),
    ("I-10", "Authorization", "CQL.copy with a Modifier only fires where() when a WHERE already exists", "Unfiltered list requests — the common case — passed through unrestricted", "Switched to Select.copy with an explicit predicate", "Closed", "2026-08-15"),
    ("I-11", "Authorization", "Select.where() replaced rather than ANDed the caller's filter", "User $filter was discarded; more rows returned than requested", "Existing predicate now combined explicitly with CQL.and", "Closed", "2026-08-15"),
    ("I-12", "Deployment", "mta.yaml had no approuter, so nothing a browser could reach", "The project had never produced a deployable archive", "Approuter added with XSUAA login and token forwarding; mbt build now produces konstryx_0.1.0.mtar (68.9 MB) with the exec jar, router, 234 HDI artifacts and xs-security.json", "Closed", "2026-08-15"),
    ("I-24", "Deployment", "No BTP target available for KONSTRYX", "The CF CLI is targeted at an unrelated client subaccount (LAND MARK INTERNATIONAL) and its token has expired. The archive is built but cannot be pushed", "Needs a KONSTRYX subaccount with Cloud Foundry, HANA Cloud and XSUAA entitlements, and a fresh cf login", "Open", "2026-08-15"),
    ("I-13", "Demo data", "Wireframe request header (716,044) does not reconcile with its own line values (685,080)", "Any stakeholder demo shows inconsistent totals", "—", "Open", "2026-08-15"),
    ("I-14", "Plan", "Personalization, attachments and approvals on every object were not in the 22-week MVP estimate", "Plan credibility; 434 screens each inherit this layer", "—", "Open", "2026-08-15"),
    ("I-15", "Authorization", "Admin persona cannot read workflow/project services", "Service-level @requires still gates entry by XSUAA role; the Admin mock user lacks those roles", "By design today, but role collections need reviewing before go-live", "Open", "2026-08-15"),
    ("I-17", "Authorization", "Enforcement broke every worklist request with HTTP 500", "The handler filters on catalogue paths project.code and company.code, but RequestOverview is a grouped projection that exposed projectCode as a flat string and dropped both associations. Missed because enforcement was verified against the service directly and the UI was not re-tested after", "View now exposes the project and company associations. Worth adding a startup check that every declared path resolves on every projection, so this fails at boot rather than per request", "Closed", "2026-08-15"),
    ("I-18", "UI", "Local FLP sandbox does not boot", "flp.html fails with 'Cannot read properties of null (reading src)'. sandbox2 is deprecated since 1.136 and the newer sap/ushell/sandbox boot contract differs", "Open — either finish the sandbox boot or verify launchpad integration directly on BTP", "Open", "2026-08-15"),
    ("I-16", "Upgrade safety", "Client configuration in db/data is reset on every upgrade", "cds build --production generates konstryx.nr-NumberRangeObject.hdbtabledata. HDI re-imports managed rows on redeploy, so a client who changes a number range scope or pattern would silently have it reverted at the next upgrade", "Fixed: configurable content moved to versioned packs applied insert-if-missing. Proven — client re-scoped RR to COMPANY, an upgrade pack containing an RR row was applied, RR kept the client's scope and only the new row arrived", "Closed", "2026-08-15"),
    ("I-19", "Content packs", "Packs applied in filename order, not version order", "number-ranges-v2.json sorted before number-ranges.json because '-' precedes '.', so 1.0.1 applied before 1.0.0 and the base pack reported its own rows as pre-existing", "Sort by packId then zero-padded version segments, so 1.0.10 follows 1.0.9", "Closed", "2026-08-15"),
    ("I-29", "Services", "ProjectService dropped associations to master data", "CAP omits an association whose target is not exposed in the same service, so a project CBS node could not say which library node it came from and a planned resource could not name its resource. ProjectResources was unusable in a UI", "Expose ResourceCatalog and CBSLibrary read-only in ProjectService for resolution; maintenance stays in MasterDataService", "Closed", "2026-08-15"),
    ("I-27", "Test data", "Seed data violated the model's own hierarchy rule", "Every resource was an L5 with no parent, so activation failed with 'L5 needs a parent at L4' and no seeded master could be edited or saved. 'Below this node' was empty on every page", "Replaced with a real L1-L5 tree of 28 nodes, leaf IDs preserved so rates and request lines still resolve", "Closed", "2026-08-15"),
    ("I-28", "UI", "Draft editing failed three ways", "draftActivate bound against the draftEdit operation context produced nested deferred bindings; expanding associations on a draft failed with 'invalid segment' and swallowed the edit-mode switch; draftEdit refused a second draft, making a master permanently uneditable after a mid-edit tab close", "Re-resolve the draft as an ordinary context, expand only the stored record, and resume an existing draft rather than erroring", "Closed", "2026-08-15"),
    ("I-25", "UI", "Blank screen after the app shell was removed", "Stripping sap.tnt.ToolPage removed the control establishing layout height. UI5 injects a UIArea div that takes no height of its own, and the NavContainer below sets overflow:hidden - so every control was clipped to nothing while still reporting a correct size. DOM text, network and per-element geometry checks all passed while the screen was empty", "Declared the full height chain in style.css. Found by painting a test div (which showed) and then walking the ancestor chain to #container-uiarea", "Closed", "2026-08-15"),
    ("I-26", "Build", "mbt build silently removes cds-dk", "The MTA build runs npm install --production at the root, pruning devDependencies. @sap/cds-dk goes with them and the next Maven build fails with 'cds' is not recognized", "Run npm install after an MTA build", "Closed", "2026-08-15"),
    ("I-23", "Data model", "Self-referencing compositions crashed draft activation", "ResourceNode, CBSNode and CBSInstance each declared children as a Composition of themselves. CAP expands compositions recursively on draft activation, so activating any hierarchy master died with a stack overflow in DraftActionsHandler. Composition also implies cascade-delete, so removing an L2 would have taken its subtree with it", "Changed to Association to many; the hierarchy is owned by parent and children are the inverse", "Closed", "2026-08-15"),
    ("I-21", "Masters", "Company-scoped authorization hid every group-scoped master", "Scoped masters declared companyPath in the authorization catalogue, so the handler added owningCompany.code = 'INFC', which is null on every GROUP row. Each steward saw only their own local master and none of the five shared ones", "The company dimension for scoped entities now belongs to MasterScopeHandler alone; the catalogue declares no companyPath for them", "Closed", "2026-08-15"),
    ("I-22", "Masters", "Promotion approval threw NullPointerException", "Map.of rejects null values and clearing owningCompany_ID is precisely a null", "Use a HashMap for update payloads that clear fields", "Closed", "2026-08-15"),
    ("I-20", "Number ranges", "A range configured on an entity without docNo would fail every create", "konstryx.wf.AdvisoryDecision has no document number; the handler would have written the field regardless", "Handler now skips targets with no docNo element, so a misconfiguration is ignored rather than breaking creates", "Closed", "2026-08-15"),
    ("I-29", "Approvals", "An approval could be raised against an object that does not exist", "The target is polymorphic - an entity name and a key, not a typed association - so no foreign key would ever object. The approval would sit in an inbox indefinitely pointing at nothing", "Submission now checks the entity is in the model and the row is really there", "Closed", "2026-08-15"),
    ("I-30", "Content", "A failing content pack left rows behind that nothing recorded as applied", "Rows were inserted one at a time and the pack's own record written last. When that record failed - a description longer than its column - the rows stayed and the pack looked unapplied", "Each pack now applies in its own change set and is abandoned whole on failure", "Closed", "2026-08-15"),
    ("I-32", "Personalization", "Every user could read and write every other user's saved layouts", "UserVariant was modelled with a user column and no handler behind it, so nothing filtered reads and nothing stopped the payload naming a different owner. A saved filter names projects, counterparties and cost codes, so another person's variant list tells you what they work on", "Owner is taken from the session; reads are narrowed to own plus published; writes to someone else's answer 404 rather than 403 so the key is not confirmed", "Closed", "2026-08-15"),
    ("I-31", "Approvals", "withdraw and delegate returned 501 Not Implemented", "Both actions were declared in CDS without a return type while their handlers set one, which CAP rejects at the protocol layer rather than at build time. Only reachable by calling them - the build was clean", "Declared returns String on both", "Closed", "2026-08-15"),
]
body(ws, 6, issues, [8, 16, 62, 50, 56, 10, 12], status_col=6)
ws.freeze_panes = "A6"

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = False

wb.save(OUT)
print("written:", OUT)
